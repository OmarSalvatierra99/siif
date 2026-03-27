"""Analyze Excel files in example_SIIF/input/OPD/ — read-only."""
import openpyxl
import pandas as pd
import os
import sys

BASE = "/home/gabo/portfolio/projects/08-siif/example_SIIF/input/OPD"

def analyze_file(filepath):
    fname = os.path.basename(filepath)
    fsize = os.path.getsize(filepath)
    print("=" * 100)
    print(f"FILE: {fname}")
    print(f"SIZE: {fsize:,} bytes ({fsize/1024:.1f} KB)")
    print("=" * 100)

    # Open with openpyxl to check merged cells and raw cell values
    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    except Exception as e:
        print(f"  ERROR opening with openpyxl: {e}")
        return

    print(f"NUMBER OF SHEETS: {len(wb.sheetnames)}")
    print(f"SHEET NAMES: {wb.sheetnames}")
    print()

    for sname in wb.sheetnames:
        ws = wb[sname]
        print("-" * 90)
        print(f"  SHEET: '{sname}'")
        print(f"  DIMENSIONS: {ws.dimensions}")
        print(f"  MAX ROW: {ws.max_row}")
        print(f"  MAX COL: {ws.max_column}")

        # Merged cells
        merged = list(ws.merged_cells.ranges)
        if merged:
            print(f"  MERGED CELLS ({len(merged)} ranges):")
            for mc in merged[:20]:
                print(f"    {mc}")
            if len(merged) > 20:
                print(f"    ... and {len(merged) - 20} more")
        else:
            print("  MERGED CELLS: None")

        # First 15 rows raw
        print()
        print(f"  FIRST 15 ROWS (raw cell values):")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=False), start=1):
            values = []
            for cell in row:
                v = cell.value
                if v is None:
                    values.append("")
                else:
                    values.append(str(v))
            print(f"    Row {row_idx:>3}: {values}")

        # Now use pandas to find data structure
        print()
        print(f"  PANDAS READ (first 20 rows, header=None):")
        try:
            df = pd.read_excel(filepath, sheet_name=sname, header=None, nrows=20)
            print(f"    Shape: {df.shape}")
            with pd.option_context('display.max_columns', None, 'display.width', 200, 'display.max_colwidth', 40):
                print(df.to_string(index=True))
        except Exception as e:
            print(f"    Error reading with pandas: {e}")

        # Detect auxiliar contable format
        print()
        auxiliar_keywords = {'fecha', 'poliza', 'póliza', 'beneficiario', 'cargo', 'abono', 'saldo'}
        found_keywords = set()
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True), start=1):
            row_lower = {str(v).lower().strip() for v in row if v is not None}
            matches = auxiliar_keywords & row_lower
            if matches:
                found_keywords |= matches
                if header_row is None:
                    header_row = row_idx

        if found_keywords:
            print(f"  AUXILIAR CONTABLE DETECTION: YES - found keywords {found_keywords} at/near row {header_row}")
        else:
            print(f"  AUXILIAR CONTABLE DETECTION: NO - none of {auxiliar_keywords} found in first 20 rows")

        # Sample data rows after detected header
        if header_row:
            print(f"  SAMPLE DATA ROWS (rows {header_row+1} to {header_row+5}):")
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=min(header_row+5, ws.max_row), values_only=True), start=header_row+1):
                print(f"    Row {row_idx}: {list(row)}")
        else:
            # Show rows 3-7 as sample anyway
            print(f"  SAMPLE DATA ROWS (rows 3-7):")
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=min(7, ws.max_row), values_only=True), start=3):
                print(f"    Row {row_idx}: {list(row)}")

        # Count non-empty rows
        total_nonempty = 0
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                total_nonempty += 1
        print(f"  NON-EMPTY ROWS: {total_nonempty}")

        # Notable patterns
        print()
        print(f"  NOTABLE PATTERNS:")
        # Check for subtotal rows
        subtotal_count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            for v in row:
                if v is not None and isinstance(v, str) and ('total' in v.lower() or 'subtotal' in v.lower()):
                    subtotal_count += 1
                    break
        if subtotal_count:
            print(f"    - Found {subtotal_count} rows containing 'total'/'subtotal'")
        if merged:
            print(f"    - Has {len(merged)} merged cell ranges")
        # Check for multi-line headers (first few rows all look like text)
        text_rows = 0
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
            vals = [v for v in row if v is not None]
            if vals and all(isinstance(v, str) for v in vals):
                text_rows += 1
        if text_rows >= 2:
            print(f"    - First {text_rows} rows are all text (possible multi-line header)")

        print()

    wb.close()


# Batch 1: Smaller files
batch1 = [
    "3 Patrimonio 1T 25.xlsx",
    "4 Ingresos 1T 25.xlsx",
    "25 2T 3 Patrimonio.xlsx",
    "25 2T 4 Ingresos.xlsx",
]

# Batch 2
batch2 = [
    "1 Activo 1T 25.xlsx",
    "2 Pasivo 1T 25.xlsx",
    "25 2T 1 Activo.xlsx",
    "25 2T 2 Pasivo.xlsx",
]

# Batch 3
batch3 = [
    "5 Gasto 1T 25.xlsx",
    "25 2T 5 Gasto.xlsx",
]

# Batch 4
batch4 = [
    "25 2T 81 Ley Ingresos.xlsx",
    "25 2T 821 PEg Aprobado.xlsx",
    "25 2T 822 PEg por Ejercer.xlsx",
]

batch_num = int(sys.argv[1]) if len(sys.argv) > 1 else 0
batches = {1: batch1, 2: batch2, 3: batch3, 4: batch4}

if batch_num == 0:
    files = batch1 + batch2 + batch3 + batch4
else:
    files = batches.get(batch_num, [])

for fname in files:
    fpath = os.path.join(BASE, fname)
    if os.path.exists(fpath):
        analyze_file(fpath)
    else:
        print(f"FILE NOT FOUND: {fname}")
    print("\n\n")
